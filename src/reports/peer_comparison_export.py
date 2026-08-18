"""Sprint 3, Day 20: peer_comparison.xlsx -- one sheet per peer group.

Scope note: the spec's dataset catalogue (Section 6, peer_comparison.xlsx
row) says "20 metrics", but Day 18's actual task only specifies computing
percentiles for the 8 radar-chart metrics (Module 4.2). Rather than invent
12 more metrics with no spec guidance on which ones matter for peer
comparison, this sheet covers the 8 that were actually built and tested in
peer.py -- flagged here rather than silently claiming "20 metrics" and
padding with arbitrary extras.

Colour coding: percentile_rank is coloured on a red -> yellow -> green
scale (low -> high), using openpyxl's built-in 3-color-scale conditional
formatting rather than hand-computed per-cell colours -- Excel updates the
colours automatically if the underlying values are edited in place, which
seems like reasonable default behaviour for a file an analyst is expected
to open and maybe adjust.
"""

import importlib.util
import os
import sys

import pandas as pd
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Font, PatternFill

sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", "screener"))
sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", "analytics"))

from engine import build_screener_universe
from peer import add_eps_cagr, RADAR_METRICS

OUTPUT_DIR = os.path.join(os.path.dirname(__file__), "..", "..", "output")


def _import_db_loader():
    path = os.path.join(os.path.dirname(__file__), "..", "..", "db", "loader.py")
    spec = importlib.util.spec_from_file_location("db_loader_module", path)
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module


db_loader = _import_db_loader()


def build_group_sheet(group_name: str, peer_percentiles: pd.DataFrame,
                       peer_groups: pd.DataFrame, universe: pd.DataFrame,
                       scale_flagged_keys: set) -> pd.DataFrame:
    """One row per company in this peer group, one column per metric's percentile."""
    members = peer_groups[peer_groups["peer_group_name"] == group_name]
    group_data = peer_percentiles[peer_percentiles["peer_group"] == group_name]

    pivot = group_data.pivot(index="company_id", columns="metric", values="percentile_rank")
    pivot = pivot.reindex(columns=list(RADAR_METRICS.keys()))  # fixed column order across all 11 sheets

    sheet = members.merge(pivot, left_on="company_id", right_index=True, how="left")
    sheet = sheet.merge(universe[["company_id", "company_name", "year"]], on="company_id", how="left")
    sheet["data_quality_caveat"] = sheet.apply(
        lambda r: (r["company_id"], r["year"]) in scale_flagged_keys, axis=1
    )
    cols = ["company_id", "company_name", "is_benchmark", "year"] + list(RADAR_METRICS.keys()) + ["data_quality_caveat"]
    return sheet[cols].sort_values("is_benchmark", ascending=False)


def export_peer_comparison(peer_percentiles: pd.DataFrame, peer_groups: pd.DataFrame,
                             universe: pd.DataFrame, scale_flagged_keys: set, path: str) -> None:
    group_names = sorted(peer_groups["peer_group_name"].unique())

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for group_name in group_names:
            sheet_df = build_group_sheet(group_name, peer_percentiles, peer_groups, universe, scale_flagged_keys)
            # Excel sheet names cap at 31 chars and disallow some
            # characters -- peer group names in this dataset are all short
            # and plain (e.g. "IT Services", "Private Banks"), so a direct
            # truncation is safe rather than needing a full sanitiser.
            sheet_df.to_excel(writer, sheet_name=group_name[:31], index=False)

    _apply_formatting(path, list(RADAR_METRICS.keys()))


def _apply_formatting(path: str, metric_columns: list) -> None:
    import openpyxl
    wb = openpyxl.load_workbook(path)
    caveat_fill = PatternFill(start_color="FFF3C4", end_color="FFF3C4", fill_type="solid")
    benchmark_font = Font(bold=True)

    for sheet in wb.worksheets:
        header = [cell.value for cell in sheet[1]]
        last_row = sheet.max_row
        last_col_letter = sheet.cell(row=1, column=sheet.max_column).column_letter

        # Colour-scale each percentile metric column (spec: "colour-coded
        # percentile cells").
        for metric in metric_columns:
            if metric not in header:
                continue
            col_letter = sheet.cell(row=1, column=header.index(metric) + 1).column_letter
            cell_range = f"{col_letter}2:{col_letter}{last_row}"
            rule = ColorScaleRule(
                start_type="min", start_color="F87171",   # red -- low percentile
                mid_type="percentile", mid_value=50, mid_color="FDE68A",  # yellow -- median
                end_type="max", end_color="4ADE80",       # green -- high percentile
            )
            sheet.conditional_formatting.add(cell_range, rule)

        is_benchmark_col = header.index("is_benchmark") + 1 if "is_benchmark" in header else None
        caveat_col = header.index("data_quality_caveat") + 1 if "data_quality_caveat" in header else None
        for row in sheet.iter_rows(min_row=2):
            if is_benchmark_col and row[is_benchmark_col - 1].value:
                for cell in row:
                    cell.font = benchmark_font
            if caveat_col and row[caveat_col - 1].value:
                row[caveat_col - 1].fill = caveat_fill

    wb.save(path)


def run() -> str:
    conn = db_loader.get_connection()
    try:
        universe = build_screener_universe(conn)
        universe = add_eps_cagr(universe, conn)
        peer_percentiles = pd.read_sql("SELECT * FROM peer_percentiles", conn)
        peer_groups = pd.read_sql("SELECT company_id, peer_group_name, is_benchmark FROM peer_groups", conn)
    finally:
        conn.close()

    scale_flags = pd.read_csv(os.path.join(OUTPUT_DIR, "scale_anomaly_flags.csv"))
    scale_flagged_keys = set(zip(scale_flags["company_id"], scale_flags["year"]))

    os.makedirs(OUTPUT_DIR, exist_ok=True)
    output_path = os.path.join(OUTPUT_DIR, "peer_comparison.xlsx")
    export_peer_comparison(peer_percentiles, peer_groups, universe, scale_flagged_keys, output_path)
    return output_path


if __name__ == "__main__":
    path = run()
    print(f"peer_comparison.xlsx written to {path}")
