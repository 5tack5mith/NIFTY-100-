"""Sprint 3 Ranking Engine -- Day 17 (+ Day 16's preset testing).

Composite score = 50% profitability + 30% growth + 20% valuation (spec
Module 3, Feature 3.3). The spec names these three weighted buckets but
doesn't say which raw metrics feed "profitability" specifically for this
composite -- so this module borrows the ONE composite formula the spec
does fully specify anywhere (Section 13's "Composite Quality Score":
0.3xROE + 0.25xFCF + 0.25xROCE + 0.20xD/E, each sub-score normalised 0-100
via P10/P90 winsorisation) and uses it as the profitability bucket. This
is a judgment call, not something copied verbatim from a single spec
passage -- but it's grounded in the one place the spec actually shows its
work on how to build a normalised composite, rather than an arbitrary
pick of "just use ROE".

Growth bucket = average of normalised 5yr Revenue CAGR and 5yr PAT CAGR.
Valuation bucket = average of normalised (P/E, P/B), INVERTED -- for
valuation, lower is better (cheaper), unlike every other metric here where
higher is better.
"""

import os
import sys

import pandas as pd

sys.path.insert(0, os.path.dirname(__file__))
from engine import build_screener_universe, run_preset, load_config, db_loader

OUTPUT_DIR = os.path.join(os.path.dirname(__file__), "..", "..", "output")


def winsorized_score(series: pd.Series, lower_is_better: bool = False) -> pd.Series:
    """Normalise a metric to 0-100 using P10/P90 winsorisation (spec 13).

    Winsorising to the 10th/90th percentile before scaling means a single
    extreme outlier (e.g. one company with a wildly high ROE from a
    one-off gain) can't compress everyone else's score toward zero --
    P10/P90 is explicitly the spec's stated method, not min/max scaling.
    """
    p10, p90 = series.quantile(0.10), series.quantile(0.90)
    if p90 == p10:
        # Every value in range is identical (or only one non-null value) --
        # no meaningful spread to score against; give everyone the same
        # midpoint rather than dividing by zero.
        return pd.Series(50.0, index=series.index).where(series.notna())
    clipped = series.clip(lower=p10, upper=p90)
    score = (clipped - p10) / (p90 - p10) * 100
    if lower_is_better:
        score = 100 - score
    return score


def compute_bucket_scores(universe: pd.DataFrame) -> pd.DataFrame:
    df = universe.copy()

    roe_score = winsorized_score(df["return_on_equity_pct"])
    fcf_score = winsorized_score(df["free_cash_flow_cr"])
    roce_score = winsorized_score(df["computed_roce_pct"])
    de_score = winsorized_score(df["debt_to_equity"], lower_is_better=True)
    # Weighted average over whatever sub-scores are actually available for
    # each row, rather than requiring all 4 -- a company missing just ROCE
    # (e.g. due to the Sprint 2 scale anomaly) shouldn't lose its entire
    # profitability score over one missing input. Weights are renormalised
    # to sum to 1 over the available components for that row.
    weights = pd.DataFrame({
        "roe": roe_score.notna() * 0.30, "fcf": fcf_score.notna() * 0.25,
        "roce": roce_score.notna() * 0.25, "de": de_score.notna() * 0.20,
    })
    weighted_sum = (
        roe_score.fillna(0) * weights["roe"] + fcf_score.fillna(0) * weights["fcf"]
        + roce_score.fillna(0) * weights["roce"] + de_score.fillna(0) * weights["de"]
    )
    weight_total = weights.sum(axis=1)
    df["profitability_score"] = (weighted_sum / weight_total).where(weight_total > 0)

    rev_cagr_score = winsorized_score(df["revenue_cagr_5yr_pct"])
    pat_cagr_score = winsorized_score(df["pat_cagr_5yr_pct"])
    df["growth_score"] = pd.concat([rev_cagr_score, pat_cagr_score], axis=1).mean(axis=1, skipna=True)

    pe_score = winsorized_score(df["pe_ratio"], lower_is_better=True)
    pb_score = winsorized_score(df["pb_ratio"], lower_is_better=True)
    df["valuation_score"] = pd.concat([pe_score, pb_score], axis=1).mean(axis=1, skipna=True)

    # A company missing an entire bucket (e.g. no growth data at all) gets
    # no composite score, rather than a composite silently computed from
    # only 1 of 3 buckets -- that would let a single available number
    # masquerade as a fair overall ranking.
    buckets = df[["profitability_score", "growth_score", "valuation_score"]]
    df["composite_score"] = (
        0.5 * buckets["profitability_score"] + 0.3 * buckets["growth_score"]
        + 0.2 * buckets["valuation_score"]
    ).where(buckets.notna().all(axis=1))

    return df


def add_rankings(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df["overall_rank"] = df["composite_score"].rank(ascending=False, method="min")
    # Sector-relative rank (spec 3.3/3.4: "RANK() OVER sector partition") --
    # ranking a small IT sector (6 companies) against the 23-company
    # Financials sector on the same absolute scale would be misleading, so
    # this is the rank an analyst filtering "top of my sector" would
    # actually want.
    df["sector_rank"] = df.groupby("broad_sector")["composite_score"].rank(ascending=False, method="min")
    return df


def export_screener_output(universe_ranked: pd.DataFrame, preset_results: dict, path: str) -> None:
    """Write screener_output.xlsx: one sheet per preset (D16) + one 'All Ranked' sheet (D17)."""
    display_cols = [
        "company_id", "company_name", "broad_sector", "year",
        "composite_score", "overall_rank", "sector_rank",
        "profitability_score", "growth_score", "valuation_score",
        "return_on_equity_pct", "computed_roce_pct", "debt_to_equity",
        "free_cash_flow_cr", "pe_ratio", "pb_ratio", "dividend_yield_pct",
        "revenue_cagr_5yr_pct", "pat_cagr_5yr_pct",
        "net_profit_margin_pct", "interest_coverage",
        "data_quality_caveat",
    ]
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        universe_ranked.sort_values("overall_rank")[display_cols].to_excel(
            writer, sheet_name="All Ranked", index=False
        )
        for preset_name, result_df in preset_results.items():
            result_df.sort_values("overall_rank")[display_cols].to_excel(
                writer, sheet_name=preset_name, index=False
            )
    _highlight_caveat_rows(path)


def _highlight_caveat_rows(path: str) -> None:
    """Colour any row with data_quality_caveat=True on every sheet, so the
    Sprint 2 scale-anomaly companies (BEL/HAL/INDIGO/LT) can't silently
    rank on numbers already flagged as unreliable -- exactly the
    requirement from the Sprint 3 kickoff instructions.
    """
    import openpyxl
    from openpyxl.styles import PatternFill

    caveat_fill = PatternFill(start_color="FFF3C4", end_color="FFF3C4", fill_type="solid")
    wb = openpyxl.load_workbook(path)
    for sheet in wb.worksheets:
        header = [cell.value for cell in sheet[1]]
        if "data_quality_caveat" not in header:
            continue
        caveat_col = header.index("data_quality_caveat") + 1
        for row in sheet.iter_rows(min_row=2):
            if row[caveat_col - 1].value:
                for cell in row:
                    cell.fill = caveat_fill
    wb.save(path)


def run() -> dict:
    """Runs D16 (test all 6 presets) and D17 (rank + export). Returns a
    summary dict so the caller (or a REPL/test) can sanity-check counts
    without re-parsing the Excel file.
    """
    conn = db_loader.get_connection()
    try:
        universe = build_screener_universe(conn)
    finally:
        conn.close()

    universe_scored = compute_bucket_scores(universe)
    universe_ranked = add_rankings(universe_scored)

    config = load_config()
    preset_results = {}
    preset_counts = {}
    for preset_name in config["presets"]:
        filtered = run_preset(universe_ranked, preset_name, config)
        preset_results[preset_name] = filtered
        preset_counts[preset_name] = {
            "matched": len(filtered),
            "with_caveat": int(filtered["data_quality_caveat"].sum()),
        }

    os.makedirs(OUTPUT_DIR, exist_ok=True)
    output_path = os.path.join(OUTPUT_DIR, "screener_output.xlsx")
    export_screener_output(universe_ranked, preset_results, output_path)

    return {
        "universe_size": len(universe_ranked),
        "preset_counts": preset_counts,
        "output_path": output_path,
    }


if __name__ == "__main__":
    summary = run()
    print(f"Screener universe: {summary['universe_size']} companies")
    print()
    print("Preset results (D16 -- verify each makes business sense):")
    for preset_name, counts in summary["preset_counts"].items():
        print(f"  {preset_name}: {counts['matched']} companies matched"
              f" ({counts['with_caveat']} carry the data-quality caveat)")
    print(f"\nWritten to {summary['output_path']}")
