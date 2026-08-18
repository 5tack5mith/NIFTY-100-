"""Sprint 4, Day 26: Valuation module -- valuation_summary.xlsx.

P/E, P/B, EV/EBITDA all come directly from market_cap.xlsx (spec 6.3: these
are pre-simulated valuation multiples, not something this project computes
from scratch). FCF Yield is the one genuinely computed metric here:
FCF / market_cap_crore x 100 (spec 13), pulling FCF from the
financial_ratios table Sprint 2 built.

"Overvaluation flags" -- the spec gives per-metric fair-value benchmarks
(P/E 15-25x fair, P/B >3x expensive, EV/EBITDA 12-18x fair) but never
defines a combined "is this company overvalued" rule. Judgment call made
here: flag a company as overvalued if AT LEAST 2 of its 3 valuation
multiples exceed their spec-stated upper "fair" bound (P/E>25, P/B>3,
EV/EBITDA>18) -- requiring 2-of-3 agreement rather than any single metric
avoids one noisy/simulated multiple alone triggering a flag, while each
individual per-metric flag is also kept as its own column so nothing is
hidden behind the combined verdict.

Since market_cap.xlsx is spec-confirmed SIMULATED data (Section 6, R-06),
this whole module is explicitly a demonstration of the valuation logic
against synthetic multiples, not real market-derived overvaluation calls --
every dashboard surface that shows these numbers carries that notice (see
data_loader.render_simulated_data_notice()).
"""

import importlib.util
import os
import sys

import pandas as pd

OUTPUT_DIR = os.path.join(os.path.dirname(__file__), "..", "..", "output")

PE_OVERVALUED_THRESHOLD = 25.0
PB_OVERVALUED_THRESHOLD = 3.0
EV_EBITDA_OVERVALUED_THRESHOLD = 18.0


def _import_db_loader():
    path = os.path.join(os.path.dirname(__file__), "..", "..", "db", "loader.py")
    spec = importlib.util.spec_from_file_location("db_loader_module", path)
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module


db_loader = _import_db_loader()


def fcf_yield(fcf_cr, market_cap_crore):
    """FCF Yield = FCF / market_cap_crore x 100 (spec 13). None if market cap missing/zero."""
    if pd.isna(fcf_cr) or pd.isna(market_cap_crore) or market_cap_crore == 0:
        return None
    return (fcf_cr / market_cap_crore) * 100


def build_valuation_summary(conn) -> pd.DataFrame:
    companies = pd.read_sql("SELECT id AS company_id, company_name FROM companies", conn)
    sectors = pd.read_sql("SELECT company_id, broad_sector FROM sectors", conn)

    mc_all = pd.read_sql(
        "SELECT company_id, year, market_cap_crore, pe_ratio, pb_ratio, ev_ebitda, dividend_yield_pct FROM market_cap",
        conn,
    )
    mc_latest = mc_all.sort_values("year").groupby("company_id", as_index=False).tail(1)

    fr_all = pd.read_sql("SELECT company_id, year, free_cash_flow_cr FROM financial_ratios", conn)
    fr_latest = fr_all.sort_values("year").groupby("company_id", as_index=False).tail(1)
    fr_latest = fr_latest.rename(columns={"year": "fr_year"})

    scale_flags = pd.read_csv(os.path.join(OUTPUT_DIR, "scale_anomaly_flags.csv"))
    flagged_keys = set(zip(scale_flags["company_id"], scale_flags["year"]))

    df = mc_latest.merge(companies, on="company_id", how="left")
    df = df.merge(sectors, on="company_id", how="left")
    df = df.merge(fr_latest, on="company_id", how="left")

    df["fcf_yield_pct"] = df.apply(lambda r: fcf_yield(r["free_cash_flow_cr"], r["market_cap_crore"]), axis=1)

    df["pe_overvalued"] = df["pe_ratio"] > PE_OVERVALUED_THRESHOLD
    df["pb_overvalued"] = df["pb_ratio"] > PB_OVERVALUED_THRESHOLD
    df["ev_ebitda_overvalued"] = df["ev_ebitda"] > EV_EBITDA_OVERVALUED_THRESHOLD
    overvalued_count = df[["pe_overvalued", "pb_overvalued", "ev_ebitda_overvalued"]].sum(axis=1)
    df["overvaluation_flag"] = overvalued_count >= 2

    # data_quality_caveat here checks the financial_ratios year (fr_year,
    # where FCF -- and therefore fcf_yield -- comes from), not the
    # market_cap year, since FCF is the metric that traces back to the
    # Sprint 2 scale anomaly (via total_assets-dependent ratios in the same
    # row), not the market_cap multiples themselves.
    df["data_quality_caveat"] = df.apply(
        lambda r: (r["company_id"], r["fr_year"]) in flagged_keys, axis=1
    )

    return df[[
        "company_id", "company_name", "broad_sector", "year",
        "market_cap_crore", "pe_ratio", "pb_ratio", "ev_ebitda", "dividend_yield_pct",
        "fcf_yield_pct", "pe_overvalued", "pb_overvalued", "ev_ebitda_overvalued",
        "overvaluation_flag", "data_quality_caveat",
    ]]


def export_valuation_summary(df: pd.DataFrame, path: str) -> None:
    df.to_excel(path, index=False, sheet_name="Valuation Summary")
    _highlight_flags(path)


def _highlight_flags(path: str) -> None:
    import openpyxl
    from openpyxl.styles import PatternFill

    overvalued_fill = PatternFill(start_color="FCA5A5", end_color="FCA5A5", fill_type="solid")
    caveat_fill = PatternFill(start_color="FFF3C4", end_color="FFF3C4", fill_type="solid")

    wb = openpyxl.load_workbook(path)
    sheet = wb["Valuation Summary"]
    header = [cell.value for cell in sheet[1]]
    overvalued_col = header.index("overvaluation_flag") + 1
    caveat_col = header.index("data_quality_caveat") + 1

    for row in sheet.iter_rows(min_row=2):
        if row[overvalued_col - 1].value:
            for cell in row:
                cell.fill = overvalued_fill
        if row[caveat_col - 1].value:
            row[caveat_col - 1].fill = caveat_fill

    wb.save(path)


def run() -> tuple:
    conn = db_loader.get_connection()
    try:
        df = build_valuation_summary(conn)
    finally:
        conn.close()

    os.makedirs(OUTPUT_DIR, exist_ok=True)
    output_path = os.path.join(OUTPUT_DIR, "valuation_summary.xlsx")
    export_valuation_summary(df, output_path)
    return output_path, df


if __name__ == "__main__":
    path, df = run()
    print(f"valuation_summary.xlsx: {len(df)} rows written to {path}")
    print(f"Overvalued (2+ metrics): {df['overvaluation_flag'].sum()}")
    print(f"Data-quality caveat rows: {df['data_quality_caveat'].sum()}")
