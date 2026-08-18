"""Sprint 5, Day 31-32: Cash Flow Intelligence module (Module 7).

Deeper cash-flow analysis than Sprint 2's ratios.py/cashflow_kpis.py --
those computed per-company-year metrics; this module summarises each
company's PATTERN over its history (5yr average CFO quality, FCF CAGR,
whether it's actively deleveraging) into one row per company.

Scale-sensitivity check (per the Sprint 5 kickoff instructions): most of
Module 7's features (CFO Quality, CapEx Intensity, FCF CAGR, FCF
Conversion, Distress Pattern) depend only on CFO/CFI/CFF/sales/
operating_profit/net_profit -- none of which are affected by the Sprint 2
BEL/HAL/INDIGO/LT balance-sheet scale issue (that issue is specifically in
total_assets/equity_capital/reserves/borrowings). The ONE exception is
7.5, Debt Repayment Detection, which explicitly compares YoY `borrowings`
-- a mis-scaled balance sheet could show a misleading "debt dropped 100x"
between a flagged and an unflagged year for the same company (this
actually happened for LT: flagged 2013-2019, clean from 2020 onward -- a
YoY borrowings comparison spanning that boundary would look like massive
debt repayment that never really happened). That rule gets the caveat
treatment; the other 6 don't.
"""

import importlib.util
import os
import sys

import pandas as pd

sys.path.insert(0, os.path.join(os.path.dirname(__file__)))
sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", "screener"))
from cagr import cagr_for_company
from cashflow_kpis import cfo_quality_score, capex_intensity, fcf_conversion_rate, classify_capital_allocation

OUTPUT_DIR = os.path.join(os.path.dirname(__file__), "..", "..", "output")


def cfo_quality_label(avg_ratio) -> str:
    """7.1: 5yr avg CFO/PAT > 1.0 = High Quality Earnings, < 0.5 = Accrual Risk."""
    if avg_ratio is None:
        return "Insufficient Data"
    if avg_ratio > 1.0:
        return "High Quality Earnings"
    if avg_ratio < 0.5:
        return "Accrual Risk"
    return "Moderate"


def capex_intensity_label(pct) -> str:
    """7.2: <3% = asset-light, >8% = capital intensive."""
    if pct is None:
        return "Insufficient Data"
    if pct < 3:
        return "Asset-Light"
    if pct > 8:
        return "Capital Intensive"
    return "Moderate"


def fcf_conversion_label(pct) -> str:
    """7.4: >60% = efficient, <30% = CapEx heavy."""
    if pct is None:
        return "Insufficient Data"
    if pct > 60:
        return "Efficient"
    if pct < 30:
        return "CapEx Heavy"
    return "Moderate"


def detect_deleveraging(cf_company: pd.DataFrame, bs_company: pd.DataFrame) -> bool:
    """7.5: latest year CFF < 0 AND borrowings declined vs the prior year.

    SCALE-SENSITIVE (uses borrowings) -- see module docstring. Caller is
    responsible for checking the caveat before trusting a True result.
    """
    if len(bs_company) < 2 or cf_company.empty:
        return False
    latest_cff = cf_company.sort_values("year")["financing_activity"].iloc[-1]
    bs_sorted = bs_company.sort_values("year")
    latest_borrowings = bs_sorted["borrowings"].iloc[-1]
    prior_borrowings = bs_sorted["borrowings"].iloc[-2]
    if pd.isna(latest_cff) or pd.isna(latest_borrowings) or pd.isna(prior_borrowings):
        return False
    return latest_cff < 0 and latest_borrowings < prior_borrowings


def detect_distress(cf_company: pd.DataFrame) -> bool:
    """7.6: latest year CFO < 0 AND CFF > 0 -- raising funds to cover an operating shortfall."""
    if cf_company.empty:
        return False
    latest = cf_company.sort_values("year").iloc[-1]
    if pd.isna(latest["operating_activity"]) or pd.isna(latest["financing_activity"]):
        return False
    return latest["operating_activity"] < 0 and latest["financing_activity"] > 0


def build_cashflow_intelligence(pl: pd.DataFrame, bs: pd.DataFrame, cf: pd.DataFrame,
                                  companies: pd.DataFrame, flagged_keys: set) -> pd.DataFrame:
    rows = []
    for company_id in companies["company_id"]:
        pl_c = pl[pl["company_id"] == company_id].sort_values("year")
        bs_c = bs[bs["company_id"] == company_id].sort_values("year")
        cf_c = cf[cf["company_id"] == company_id].sort_values("year")
        if pl_c.empty or cf_c.empty:
            continue

        latest_year = pl_c["year"].iloc[-1]

        # 7.1 CFO Quality Score -- 5yr average, not a single year, per spec.
        recent_pl = pl_c.tail(5).set_index("year")
        recent_cf = cf_c.tail(5).set_index("year")
        ratios = [
            cfo_quality_score(recent_cf.loc[y, "operating_activity"], recent_pl.loc[y, "net_profit"])
            for y in recent_pl.index if y in recent_cf.index
        ]
        ratios = [r for r in ratios if r is not None]
        avg_cfo_quality = sum(ratios) / len(ratios) if ratios else None

        # 7.2 CapEx Intensity -- latest year.
        latest_cf_row = cf_c.iloc[-1]
        latest_pl_row = pl_c.iloc[-1]
        capex_pct = capex_intensity(latest_cf_row["investing_activity"], latest_pl_row["sales"])

        # 7.3 FCF CAGR -- build a per-year FCF series first (FCF isn't a
        # source column; it's operating_activity + investing_activity),
        # then reuse the same CAGR engine every other module uses.
        cf_indexed = cf_c.set_index("year")
        fcf_series = cf_indexed["operating_activity"] + cf_indexed["investing_activity"]
        fcf_cagr_result = cagr_for_company(fcf_series, windows=(5, 10))

        # 7.4 FCF Conversion -- latest year, FCF / operating_profit (EBITDA proxy per spec 5.2).
        latest_fcf = fcf_series.get(latest_year)
        conversion_pct = fcf_conversion_rate(latest_fcf, latest_pl_row["operating_profit"])

        # 7.5 Debt Repayment Detection -- SCALE-SENSITIVE.
        deleveraging = detect_deleveraging(cf_c, bs_c)
        deleveraging_caveat = deleveraging and (company_id, latest_year) in flagged_keys

        # 7.6 Distress Pattern -- not scale-sensitive.
        distress = detect_distress(cf_c)

        # 7.7 Capital Allocation Matrix -- reuse Sprint 2's classifier
        # directly rather than re-deriving the same 8-way logic here.
        latest_signs = classify_capital_allocation(
            latest_cf_row["operating_activity"], latest_cf_row["investing_activity"],
            latest_cf_row["financing_activity"], latest_pl_row["net_profit"],
        )

        rows.append({
            "company_id": company_id,
            "year": latest_year,
            "cfo_quality_score_5yr_avg": avg_cfo_quality,
            "cfo_quality_label": cfo_quality_label(avg_cfo_quality),
            "capex_intensity_pct": capex_pct,
            "capex_intensity_label": capex_intensity_label(capex_pct),
            "fcf_cagr_5yr_pct": fcf_cagr_result["cagr_5yr_pct"],
            "fcf_cagr_10yr_pct": fcf_cagr_result["cagr_10yr_pct"],
            "fcf_cagr_5yr_turnaround": fcf_cagr_result["turnaround_5yr"],
            "fcf_cagr_10yr_turnaround": fcf_cagr_result["turnaround_10yr"],
            "fcf_conversion_pct": conversion_pct,
            "fcf_conversion_label": fcf_conversion_label(conversion_pct),
            "deleveraging_flag": deleveraging,
            "deleveraging_data_quality_caveat": deleveraging_caveat,
            "distress_flag": distress,
            "capital_allocation_pattern": latest_signs,
        })

    return pd.DataFrame(rows)


def export(df: pd.DataFrame) -> tuple:
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    xlsx_path = os.path.join(OUTPUT_DIR, "cashflow_intelligence.xlsx")
    df.to_excel(xlsx_path, index=False, sheet_name="Cash Flow Intelligence")

    # distress_alerts.csv -- Module 7.6's own named output, a filtered
    # view for whoever just wants the alert list without opening the full
    # 92-row workbook.
    distress_df = df[df["distress_flag"]][["company_id", "year", "cfo_quality_score_5yr_avg", "capital_allocation_pattern"]]
    distress_path = os.path.join(OUTPUT_DIR, "distress_alerts.csv")
    distress_df.to_csv(distress_path, index=False)

    _highlight(xlsx_path)
    return xlsx_path, distress_path


def _highlight(xlsx_path: str) -> None:
    import openpyxl
    from openpyxl.styles import PatternFill

    distress_fill = PatternFill(start_color="FCA5A5", end_color="FCA5A5", fill_type="solid")
    caveat_fill = PatternFill(start_color="FFF3C4", end_color="FFF3C4", fill_type="solid")

    wb = openpyxl.load_workbook(xlsx_path)
    sheet = wb["Cash Flow Intelligence"]
    header = [cell.value for cell in sheet[1]]
    distress_col = header.index("distress_flag") + 1
    caveat_col = header.index("deleveraging_data_quality_caveat") + 1

    for row in sheet.iter_rows(min_row=2):
        if row[distress_col - 1].value:
            for cell in row:
                cell.fill = distress_fill
        if row[caveat_col - 1].value:
            row[caveat_col - 1].fill = caveat_fill

    wb.save(xlsx_path)


def run() -> tuple:
    import sqlite3
    conn = sqlite3.connect(os.path.join(os.path.dirname(__file__), "..", "..", "data", "nifty100.db"))
    try:
        pl = pd.read_sql("SELECT * FROM profitandloss", conn)
        bs = pd.read_sql("SELECT * FROM balancesheet", conn)
        cf = pd.read_sql("SELECT * FROM cashflow", conn)
        companies = pd.read_sql("SELECT id AS company_id FROM companies", conn)
    finally:
        conn.close()

    scale_flags = pd.read_csv(os.path.join(OUTPUT_DIR, "scale_anomaly_flags.csv"))
    flagged_keys = set(zip(scale_flags["company_id"], scale_flags["year"]))

    df = build_cashflow_intelligence(pl, bs, cf, companies, flagged_keys)
    xlsx_path, distress_path = export(df)
    return df, xlsx_path, distress_path


if __name__ == "__main__":
    df, xlsx_path, distress_path = run()
    print(f"cashflow_intelligence.xlsx: {len(df)} rows -> {xlsx_path}")
    print(f"distress_alerts.csv: {df['distress_flag'].sum()} companies flagged -> {distress_path}")
    print(f"Deleveraging flags with data-quality caveat: {df['deleveraging_data_quality_caveat'].sum()}")
    print(df["cfo_quality_label"].value_counts())
